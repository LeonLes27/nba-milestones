# collect.py
import os
import re
import time
import json
import pandas as pd
from difflib import get_close_matches
from requests.exceptions import ReadTimeout

from nba_api.stats.static import players as static_players
from nba_api.stats.endpoints import playergamelog, commonplayerinfo

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ============================================================
# Einstellungen
# ============================================================
SEASON = "2025-26"
LAST_N = 8

# etwas höher, damit wir NBA nicht nerven
SLEEP_BETWEEN_CALLS = 0.25

# 0 = still, 1 = warn, 2 = info
LOG_LEVEL = 0

# erkennen, ob wir im GitHub-Runner sind
RUNNING_IN_CI = os.environ.get("GITHUB_ACTIONS") == "true"
# damit GitHub nicht 40 Minuten läuft:
CI_PLAYER_LIMIT = 120  # kannst du später auf 200 oder 300 setzen

# Basis-Pfad (Projektwurzel)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

input_datei = os.path.join(BASE_DIR, "PlayerNames.csv")
output_xlsx = os.path.join(BASE_DIR, "TeamStatistiken_Meilensteine.xlsx")
not_found_csv = os.path.join(BASE_DIR, "not_found_names.csv")

# ============================================================
# Logging
# ============================================================
def log_info(msg):
    if LOG_LEVEL >= 2:
        print(msg)

def log_warn(msg):
    if LOG_LEVEL >= 1:
        print(msg)

def log_error(msg):
    print(msg)

# ============================================================
# Retry – für GitHub kürzer machen
# ============================================================
def retry_api_call(callable_fn, retries=1, delay=2, on_timeout_msg=None):
    """
    auf GitHub lieber schnell abbrechen statt minutenlang warten
    """
    for i in range(retries):
        try:
            return callable_fn()
        except ReadTimeout:
            if LOG_LEVEL >= 1:
                print(f"{on_timeout_msg or 'Timeout'} – warte {delay}s... (Versuch {i+1}/{retries})")
            time.sleep(delay)
    # letzte Chance fehlgeschlagen
    raise ReadTimeout("API-Anfrage mehrfach fehlgeschlagen.")

# ============================================================
# Namen normalisieren
# ============================================================
def _normalize_name(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("„", " ").replace("“", " ").replace("”", " ").replace("’", "'").replace("´", "'")
    s = s.replace('"', " ")
    s = re.sub(r"[.,;:()]", " ", s)
    s = s.replace("-", " ")
    s = re.sub(r"[\u200b\u200c\u200d]", "", s)
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

# alle offiziellen Spieler
_ALL_PLAYERS = static_players.get_players()
_name_index = {}
for p in _ALL_PLAYERS:
    norm = _normalize_name(p["full_name"])
    _name_index.setdefault(norm, []).append(p)

# Aliase
ALIAS_OVERRIDES = {
    "vj edgecomb": "vj edgecombe",
    "vj edgecome": "vj edgecombe",
    "v j edgecombe": "vj edgecombe",
    "valdez drexel v j edgecombe": "vj edgecombe",
    "vit krejci": "vít krejčí",
    "jakob poeltl": "jakob pöltl",
    "lester quiñones": "lester quinones",
    "pacome dadiet": "pacôme dadiet",
    "monte morris": "monté morris",
    "taze moore": "tazé moore",
    "isiah crawford": "isaiah crawford",
    "hood schifino jalen": "jalen hood schifino",
    "hood-schifino jalen": "jalen hood schifino",
    "dariq miller whitehead": "dariq whitehead",
    "dariq miller-whitehead": "dariq whitehead",
}

# falls jemand noch nicht in nba_api auftaucht
MANUAL_PLAYER_IDS = {
    # "vj edgecombe": 123456,
}

def resolve_player_id(name: str):
    norm = _normalize_name(name)

    # Alias anwenden
    if norm in ALIAS_OVERRIDES:
        norm = _normalize_name(ALIAS_OVERRIDES[norm])

    # manuell
    if norm in MANUAL_PLAYER_IDS:
        pid = MANUAL_PLAYER_IDS[norm]
        return pid, name

    # exakter Treffer
    candidates = _name_index.get(norm, [])
    if candidates:
        active = [c for c in candidates if c.get("is_active")]
        pick = active[0] if active else candidates[0]
        return pick["id"], pick["full_name"]

    # fuzzy
    close = get_close_matches(norm, list(_name_index.keys()), n=1, cutoff=0.82)
    if close:
        candidates = _name_index[close[0]]
        active = [c for c in candidates if c.get("is_active")]
        pick = active[0] if active else candidates[0]
        log_info(f"[Hinweis] '{name}' -> '{pick['full_name']}'")
        return pick["id"], pick["full_name"]

    log_warn(f"[Warnung] Spieler nicht gefunden: {name}")
    return None, None

current_team_cache = {}

def get_current_team_abbrev(player_id, spiele_df=None):
    if player_id in current_team_cache:
        return current_team_cache[player_id]

    team_abbr = ""
    try:
        info = retry_api_call(lambda: commonplayerinfo.CommonPlayerInfo(player_id=player_id))
        info_df = info.get_data_frames()[0]
        if "TEAM_ABBREVIATION" in info_df.columns and not info_df.empty:
            team_abbr = str(info_df.at[0, "TEAM_ABBREVIATION"]).strip()
    except Exception:
        team_abbr = ""

    if not team_abbr and spiele_df is not None and not spiele_df.empty:
        if "TEAM_ABBREVIATION" in spiele_df.columns:
            team_abbr = str(spiele_df.iloc[0]["TEAM_ABBREVIATION"]).strip()

    team_abbr = team_abbr or "FA"
    current_team_cache[player_id] = team_abbr
    return team_abbr

# ============================================================
# Meilensteine
# ============================================================
def count_milestones(spiele: pd.DataFrame, thresholds: dict) -> dict:
    counts = {}
    for metric, limits in thresholds.items():
        serie = spiele.get("STL", 0) + spiele.get("BLK", 0) if metric == "STL+BLK" else spiele.get(metric, 0)
        if not isinstance(serie, pd.Series):
            serie = pd.Series([0] * len(spiele))
        counts[metric] = {f"{limit}+": int((serie >= limit).sum()) for limit in limits}
    return counts

milestones = {
    "PTS":     [3, 5, 7, 10, 12, 14, 17, 20],
    "REB":     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    "AST":     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    "STL":     [1, 2],
    "BLK":     [1, 2],
    "STL+BLK": [1, 2, 3],
    "FG3M":    [1, 2, 3],
    "FG2M":    [2, 4, 6, 8, 10, 12],
    "FTM":     [2, 4, 6, 8, 10],
    "TO":      [1, 2, 3, 4, 5],
}

# ============================================================
# CSV sicherstellen
# ============================================================
if not os.path.exists(input_datei):
    print("🔄 PlayerNames.csv nicht gefunden – erstelle aus aktiven NBA-Spielern...")
    active_players = static_players.get_active_players()
    df_players = pd.DataFrame({"Player": [p["full_name"] for p in active_players]})
    df_players.to_csv(input_datei, index=False, sep=";")
    print(f"✅ Neue PlayerNames.csv mit {len(df_players)} Spielern erstellt.")

# ============================================================
# Spieler aus CSV laden
# ============================================================
spieler_df = pd.read_csv(input_datei, delimiter=";", quotechar='"')
spieler_df.columns = spieler_df.columns.str.strip()
if "Player" not in spieler_df.columns:
    raise ValueError("In der CSV muss eine Spalte 'Player' stehen.")
spieler_namen = spieler_df["Player"].astype(str).tolist()

# falls wir im CI laufen: abkürzen
if RUNNING_IN_CI and len(spieler_namen) > CI_PLAYER_LIMIT:
    spieler_namen = spieler_namen[:CI_PLAYER_LIMIT]
    print(f"🟡 CI-Modus: verarbeite nur die ersten {CI_PLAYER_LIMIT} Spieler.")

# ============================================================
# Hauptlogik
# ============================================================
ergebnisse = {}
not_found = []
processed_players = 0

for raw_name in spieler_namen:
    name = raw_name.strip()
    try:
        spieler_id, resolved_name = resolve_player_id(name)
        if not spieler_id:
            not_found.append(name)
            continue

        gamelog = retry_api_call(
            lambda: playergamelog.PlayerGameLog(player_id=spieler_id, season=SEASON),
            on_timeout_msg=f"PlayerGameLog {resolved_name}"
        )

        spiele = gamelog.get_data_frames()[0] if gamelog.get_data_frames() else pd.DataFrame()

        if "TOV" in spiele.columns and "TO" not in spiele.columns:
            spiele.rename(columns={"TOV": "TO"}, inplace=True)

        needed_cols = ["PTS","REB","AST","STL","BLK","FG3M","TO","FGM","FG3A","FTM","TEAM_ABBREVIATION"]
        for col in needed_cols:
            if col not in spiele.columns:
                spiele[col] = 0

        # 2er gemacht
        spiele["FG2M"] = (spiele.get("FGM", 0) - spiele.get("FG3M", 0)).clip(lower=0)

        team_abbr = get_current_team_abbrev(spieler_id, spiele_df=spiele)

        letzte_n_spiele = spiele.head(LAST_N) if not spiele.empty else spiele
        ganze_saison = spiele

        letzte_n_counts = count_milestones(letzte_n_spiele.copy(), milestones)
        ganze_saison_counts = count_milestones(ganze_saison.copy(), milestones)

        ergebnisse.setdefault(team_abbr, []).append({
            "Player": resolved_name,
            "Last N Games": letzte_n_counts,
            "Full Season": ganze_saison_counts,
            "Games Played": int(len(ganze_saison))
        })

        processed_players += 1
        time.sleep(SLEEP_BETWEEN_CALLS)

    except Exception as e:
        log_error(f"Fehler bei {name}: {e}")

# nicht gefundene Namen speichern
if not_found:
    pd.DataFrame({"not_found": not_found}).to_csv(not_found_csv, index=False)

# ============================================================
# Excel schreiben
# ============================================================
wb = Workbook()
wb.remove(wb.active)

LIGHT_GREEN  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
STRONG_GREEN = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

for team, stats in ergebnisse.items():
    ws = wb.create_sheet(title=(team or "FA")[:31])
    header = ["Milestones"]
    for s in stats:
        header += [s["Player"], f"__helper_{s['Player']}"]
    ws.append(header)
    n_players = len(stats)

    for cat, limits in milestones.items():
        for limit in limits:
            label = f"{cat} {limit}+"
            row = [label]
            helpers = []
            for j, s in enumerate(stats):
                last_val = s["Last N Games"].get(cat, {}).get(f"{limit}+", 0)
                last_pct = (last_val / LAST_N * 100) if LAST_N > 0 else 0.0

                full_val = s["Full Season"].get(cat, {}).get(f"{limit}+", 0)
                gp = s["Games Played"]
                full_pct = (full_val / gp * 100) if gp > 0 else 0.0

                row += [f"{last_val} ({last_pct:.2f}%) / {full_val} ({full_pct:.2f}%)", None]
                helpers.append((3 + 2*j, full_pct / 100.0))
            ws.append(row)
            r = ws.max_row
            for col_idx, v in helpers:
                ws.cell(row=r, column=col_idx, value=v)

    # Hilfsspalten verstecken + bedingte Formatierung
    start_row = 2
    end_row = ws.max_row
    for j in range(n_players):
        vis_col_idx = 2 + 2*j
        helper_col_idx = 3 + 2*j
        vis_letter = get_column_letter(vis_col_idx)
        helper_letter = get_column_letter(helper_col_idx)

        ws.column_dimensions[helper_letter].hidden = True

        rng = f"{vis_letter}{start_row}:{vis_letter}{end_row}"
        rule_strong = FormulaRule(formula=[f"=${helper_letter}{start_row}>=1"], fill=STRONG_GREEN)
        rule_light  = FormulaRule(formula=[f"=AND(${helper_letter}{start_row}>=0.85, ${helper_letter}{start_row}<1)"], fill=LIGHT_GREEN)
        ws.conditional_formatting.add(rng, rule_strong)
        ws.conditional_formatting.add(rng, rule_light)

wb.save(output_xlsx)
print(f"✅ Excel gespeichert: {output_xlsx}")

# ============================================================
# JSON für Web
# ============================================================
public_data_dir = os.path.join(BASE_DIR, "public", "data")
os.makedirs(public_data_dir, exist_ok=True)

json_payload = {}
for team, stats in ergebnisse.items():
    json_payload[team] = [
        {
            "player": s["Player"],
            "gp": s["Games Played"],
            "lastN": s["Last N Games"],
            "season": s["Full Season"],
        }
        for s in stats
    ]

json_path = os.path.join(public_data_dir, "milestones.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(json_payload, f, ensure_ascii=False)

print(f"🌐 JSON geschrieben -> {json_path}")
print(f"✅ Fertig – {processed_players} Spieler verarbeitet, {len(ergebnisse)} Teams.")
if RUNNING_IN_CI:
    print("ℹ️ Hinweis: CI-Modus war aktiv, deshalb nur begrenzte Spielerzahl.")
