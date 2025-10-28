# collect_local.py
import re
import time
import pandas as pd
from difflib import get_close_matches
from requests.exceptions import ReadTimeout

from nba_api.stats.static import players as static_players
from nba_api.stats.endpoints import playergamelog, commonplayerinfo

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# =========================
# Parameter
# =========================
SEASON = "2025-26"
LAST_N = 8
SLEEP_BETWEEN_CALLS = 0.10   # ggf. 0.05–0.2 testen; zu niedrig -> Rate-Limits möglich
LOG_LEVEL = 0                # 0=still, 1=warn, 2=info

# Eingabe/Output
input_datei = r"C:\Users\bruns\Desktop\NBA\NBA Data\PlayerNames.csv"   # CSV mit Spalte "Player"
output_xlsx = r"C:\Users\bruns\Desktop\NBA\NBA Data\TeamStatistiken_Meilensteine.xlsx"
not_found_csv = r"C:\Users\bruns\Desktop\NBA\NBA Data\not_found_names.csv"

# =========================
# Logging-Helper
# =========================
def log_info(msg):
    if LOG_LEVEL >= 2:
        print(msg)

def log_warn(msg):
    if LOG_LEVEL >= 1:
        print(msg)

def log_error(msg):
    print(msg)

# =========================
# Retry
# =========================
def retry_api_call(callable_fn, retries=3, delay=5, on_timeout_msg=None):
    for i in range(retries):
        try:
            return callable_fn()
        except ReadTimeout:
            if LOG_LEVEL >= 1:
                print(f"{on_timeout_msg or 'Timeout'}. Warte {delay}s... (Versuch {i+1}/{retries})")
            time.sleep(delay)
    raise ReadTimeout("API-Anfrage mehrfach fehlgeschlagen.")

# =========================
# Namen normalisieren & Index
# =========================
def _normalize_name(s: str) -> str:
    """
    Robust gegen: Sonderzeichen, typogr. Anführungszeichen, Punkte in Initialen,
    Bindestriche, Mehrfachspaces, Suffixe (Jr./Sr./II/III/IV/V).
    """
    s = (s or "").strip().lower()
    s = s.replace("„", " ").replace("“", " ").replace("”", " ").replace("‚", " ").replace("’", "'").replace("´", "'")
    s = s.replace('"', " ")
    s = re.sub(r"[.,;:()]", " ", s)   # Satzzeichen raus
    s = s.replace("-", " ")
    s = re.sub(r"[\u200b\u200c\u200d]", "", s)  # zero-width chars
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", s)  # Suffixe
    s = re.sub(r"\s+", " ", s).strip()
    return s

_ALL_PLAYERS = static_players.get_players()  # [{'id','full_name','is_active'}, ...]
_name_index = {}
for p in _ALL_PLAYERS:
    norm = _normalize_name(p["full_name"])
    _name_index.setdefault(norm, []).append(p)

# Häufige Aliase & Schreibweisen (Key: bereits normalisiert!)
ALIAS_OVERRIDES = {
    # VJ Edgecombe – diverse Varianten auf einen Nenner
    "vj edgecomb": "vj edgecombe",
    "vj edgecome": "vj edgecombe",
    "vj edgecom":  "vj edgecombe",
    "v j edgecombe": "vj edgecombe",
    "valdez drexel v j edgecombe": "vj edgecombe",

    # Diakritika / Varianten:
    "vit krejci": "vít krejčí",
    "jakob poeltl": "jakob pöltl",
    "lester quiñones": "lester quinones",
    "pacome dadiet": "pacôme dadiet",
    "monte morris": "monté morris",
    "taze moore": "tazé moore",
    "isiah crawford": "isaiah crawford",

    # Bindestrich-/Token-Varianten:
    "hood schifino jalen": "jalen hood schifino",
    "hood schifino": "jalen hood schifino",
    "hood-schifino jalen": "jalen hood schifino",
    "dariq miller whitehead": "dariq whitehead",
    "dariq miller-whitehead": "dariq whitehead",
}

# Optional: manuelle IDs (falls ein Spieler (noch) nicht in get_players() gelistet ist)
# Key MUSS normalisiert sein (wie _normalize_name es erzeugt)
MANUAL_PLAYER_IDS = {
    # Beispiel: "vj edgecombe": 123456,
}

def resolve_player_id(name: str):
    """Rückgabe: (player_id, resolved_full_name) oder (None, None)."""
    norm = _normalize_name(name)

    # 1) Alias
    if norm in ALIAS_OVERRIDES:
        norm = _normalize_name(ALIAS_OVERRIDES[norm])

    # 1a) manueller Fallback
    if norm in MANUAL_PLAYER_IDS:
        pid = MANUAL_PLAYER_IDS[norm]
        try:
            info = retry_api_call(lambda: commonplayerinfo.CommonPlayerInfo(player_id=pid),
                                  on_timeout_msg=f"CommonPlayerInfo (manual) {pid}")
            df = info.get_data_frames()[0]
            if "DISPLAY_FIRST_LAST" in df.columns and not df.empty:
                full = str(df.at[0, "DISPLAY_FIRST_LAST"]).strip()
            else:
                full = name
        except Exception:
            full = name
        return pid, full

    # 2) Exakter Treffer (normalisiert)
    candidates = _name_index.get(norm, [])
    if candidates:
        active = [c for c in candidates if c.get("is_active")]
        pick = active[0] if active else candidates[0]
        return pick["id"], pick["full_name"]

    # 3) Fuzzy
    all_norm_names = list(_name_index.keys())
    close = get_close_matches(norm, all_norm_names, n=1, cutoff=0.82)
    if close:
        candidates = _name_index[close[0]]
        active = [c for c in candidates if c.get("is_active")]
        pick = active[0] if active else candidates[0]
        log_info(f"[Hinweis] '{name}' -> interpretiert als '{pick['full_name']}'")
        return pick["id"], pick["full_name"]

    # 4) Nachname-Fallback
    tokens = norm.split()
    if tokens:
        last = tokens[-1]
        pool = [p for p in _ALL_PLAYERS if _normalize_name(p["full_name"]).endswith(" " + last)]
        if pool:
            active = [c for c in pool if c.get("is_active")]
            pick = active[0] if active else pool[0]
            log_info(f"[Hinweis] '{name}' -> (Nachname) '{pick['full_name']}'")
            return pick["id"], pick["full_name"]

    log_warn(f"[Warnung] Spieler nicht gefunden: {name}")
    return None, None

current_team_cache = {}  # player_id -> TEAM_ABBREVIATION

def get_current_team_abbrev(player_id, spiele_df=None):
    """CommonPlayerInfo → Fallback jüngstes Spiel → 'FA'."""
    if player_id in current_team_cache:
        return current_team_cache[player_id]

    team_abbr = ""
    try:
        info = retry_api_call(
            lambda: commonplayerinfo.CommonPlayerInfo(player_id=player_id),
            on_timeout_msg=f"CommonPlayerInfo für {player_id}"
        )
        info_df = info.get_data_frames()[0]
        if "TEAM_ABBREVIATION" in info_df.columns and not info_df.empty:
            team_abbr = str(info_df.at[0, "TEAM_ABBREVIATION"]).strip()
    except Exception:
        team_abbr = ""

    if (not team_abbr) and spiele_df is not None and not spiele_df.empty:
        if "TEAM_ABBREVIATION" in spiele_df.columns:
            team_abbr = str(spiele_df.iloc[0]["TEAM_ABBREVIATION"]).strip()

    team_abbr = team_abbr if team_abbr else "FA"
    current_team_cache[player_id] = team_abbr
    return team_abbr

def count_milestones(spiele: pd.DataFrame, thresholds: dict) -> dict:
    """Zählt pro Metrik, wie oft Limits (>=) erreicht wurden."""
    counts = {}
    for metric, limits in thresholds.items():
        if metric == "STL+BLK":
            serie = spiele.get("STL", 0) + spiele.get("BLK", 0)
        else:
            serie = spiele.get(metric, 0)
        if not isinstance(serie, pd.Series):
            serie = pd.Series([0] * len(spiele))
        counts[metric] = {f"{limit}+": int((serie >= limit).sum()) for limit in limits}
    return counts

# =========================
# Meilensteine (inkl. 2P & FT made)
# =========================
milestones = {
    "PTS":     [3, 5, 7, 10, 12, 14, 17, 20],
    "REB":     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    "AST":     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    "STL":     [1, 2],
    "BLK":     [1, 2],
    "STL+BLK": [1, 2, 3],
    "FG3M":    [1, 2, 3],
    "FG2M":    [2, 4, 6, 8, 10, 12],  # 2er gemacht
    "FTM":     [2, 4, 6, 8, 10],      # Freiwürfe gemacht
    "TO":      [1, 2, 3, 4, 5]
}

# =========================
# Daten einlesen
# =========================
spieler_df = pd.read_csv(input_datei, delimiter=';', quotechar='"')
spieler_df.columns = spieler_df.columns.str.strip()
if "Player" not in spieler_df.columns:
    raise ValueError("In der CSV muss eine Spalte 'Player' existieren.")
spieler_namen = spieler_df["Player"].astype(str).tolist()

# =========================
# Hauptlogik
# =========================
ergebnisse = {}
spieler_cache = {}  # name -> player_id
not_found = []

for raw_name in spieler_namen:
    name = raw_name.strip()
    try:
        if name in spieler_cache:
            spieler_id = spieler_cache[name]
            resolved_name = name
        else:
            spieler_id, resolved_name = resolve_player_id(name)
            if not spieler_id:
                not_found.append(name)
                continue
            spieler_cache[name] = spieler_id

        gamelog = retry_api_call(
            lambda: playergamelog.PlayerGameLog(player_id=spieler_id, season=SEASON),
            on_timeout_msg=f"PlayerGameLog {resolved_name} ({SEASON})"
        )
        spiele = gamelog.get_data_frames()[0] if len(gamelog.get_data_frames()) else pd.DataFrame()

        # Spalten angleichen/ergänzen
        if "TOV" in spiele.columns and "TO" not in spiele.columns:
            spiele.rename(columns={"TOV": "TO"}, inplace=True)

        needed_cols = [
            "PTS","REB","AST","STL","BLK","FG3M","TO",
            "FGM","FGA","FG3A","FTM","FTA","TEAM_ABBREVIATION"
        ]
        for col in needed_cols:
            if col not in spiele.columns:
                spiele[col] = 0

        # 2-Punkte-Würfe (gemacht)
        spiele["FG2M"] = (spiele["FGM"] - spiele["FG3M"]).clip(lower=0)

        # Team bestimmen (auch ohne Gamelog)
        team_abbr = get_current_team_abbrev(spieler_id, spiele_df=spiele)

        # Teilmengen
        letzte_n_spiele = spiele.head(LAST_N) if not spiele.empty else spiele

        ganze_saison = spiele

        # Meilensteine
        letzte_n_counts = count_milestones(letzte_n_spiele.copy(), milestones)
        ganze_saison_counts = count_milestones(ganze_saison.copy(), milestones)

        if team_abbr not in ergebnisse:
            ergebnisse[team_abbr] = []

        ergebnisse[team_abbr].append({
            "Player": resolved_name,
            "Last N Games": letzte_n_counts,
            "Full Season": ganze_saison_counts,
            "Games Played": int(len(ganze_saison))
        })

        time.sleep(SLEEP_BETWEEN_CALLS)

    except Exception as e:
        log_error(f"Fehler bei {name}: {e}")

# Unerkannte Namen in Datei schreiben (keine Konsole fluten)
if not_found:
    try:
        pd.DataFrame({"not_found": not_found}).to_csv(not_found_csv, index=False)
    except Exception as e:
        log_warn(f"Konnte not_found_names.csv nicht schreiben: {e}")

# =========================
# Excel schreiben – schnell (Hilfsspalten + echte Conditional Formatting)
# =========================
wb = Workbook()
# Startblatt entfernen
std_ws = wb.active
wb.remove(std_ws)

# Farben
LIGHT_GREEN  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
STRONG_GREEN = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

for team, stats in ergebnisse.items():
    sheet_name = (team or "FA")[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Kopfzeile mit sichtbaren Spieler-Spalten UND unsichtbaren Hilfsspalten (rechts daneben)
    header = ["Milestones"]
    for s in stats:
        header += [s["Player"], f"__helper_{s['Player']}"]  # Helper-Spalte direkt daneben
    ws.append(header)

    n_players = len(stats)

    # Datenzeilen
    for category, limits in milestones.items():
        for milestone in limits:
            label = f"{category} {milestone}+"
            row_display = [label]   # sichtbare Zellen
            helper_targets = []     # (col_idx, season_fraction)

            for j, s in enumerate(stats):
                last_n_val = s["Last N Games"].get(category, {}).get(f"{milestone}+", 0)
                last_n_pct = (last_n_val / LAST_N * 100) if LAST_N > 0 else 0.0

                full_val = s["Full Season"].get(category, {}).get(f"{milestone}+", 0)
                gp = s["Games Played"]
                full_pct = (full_val / gp * 100) if gp > 0 else 0.0

                row_display += [f"{last_n_val} ({last_n_pct:.2f}%) / {full_val} ({full_pct:.2f}%)", None]
                helper_col = 3 + 2*j  # 1-based Index der Helper-Spalte
                helper_targets.append((helper_col, full_pct / 100.0))  # 0..1 Zahl

            ws.append(row_display)
            r = ws.max_row
            for (col_idx, val) in helper_targets:
                ws.cell(row=r, column=col_idx, value=val)

    # Hilfsspalten verstecken
    for j in range(n_players):
        helper_col_idx = 3 + 2*j
        ws.column_dimensions[get_column_letter(helper_col_idx)].hidden = True

    # Conditional Formatting: >=100% starkes Grün, >=85% leichtes Grün (Basis: Helper-Spalte)
    start_row = 2
    end_row = ws.max_row
    for j in range(n_players):
        vis_col_idx    = 2 + 2*j
        helper_col_idx = 3 + 2*j
        vis_col_letter    = get_column_letter(vis_col_idx)
        helper_col_letter = get_column_letter(helper_col_idx)

        rng = f"{vis_col_letter}{start_row}:{vis_col_letter}{end_row}"

        # Formel referenziert Helper-Spalte in derselben Zeile (Zeile relativ über Startzelle)
        rule_strong = FormulaRule(formula=[f"=${helper_col_letter}{start_row}>=1"], fill=STRONG_GREEN)
        rule_light  = FormulaRule(formula=[f"=AND(${helper_col_letter}{start_row}>=0.85, ${helper_col_letter}{start_row}<1)"], fill=LIGHT_GREEN)

        ws.conditional_formatting.add(rng, rule_strong)
        ws.conditional_formatting.add(rng, rule_light)

wb.save(output_xlsx)
# --- JSON-Export für Web (für GitHub Pages) ---
import json, os

# kompaktes, frontend-freundliches Format:
# { "PHI": [ {"player":"Joel Embiid","gp":2,"lastN":{...}, "season":{...}}, ... ], ... }
json_payload = {}
for team, stats in ergebnisse.items():
    team_list = []
    for s in stats:
        team_list.append({
            "player": s["Player"],
            "gp": s["Games Played"],
            "lastN": s["Last N Games"],
            "season": s["Full Season"]
        })
    json_payload[team] = team_list

os.makedirs("public/data", exist_ok=True)
with open("public/data/milestones.json", "w", encoding="utf-8") as f:
    json.dump(json_payload, f, ensure_ascii=False)
print("🌐 JSON geschrieben -> public/data/milestones.json")
